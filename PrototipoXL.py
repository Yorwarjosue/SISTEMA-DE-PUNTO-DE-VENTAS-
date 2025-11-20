# =====================================================================================
# 1. IMPORTACIONES DE LIBRERÍAS
# =====================================================================================
import tkinter as tk
from tkinter import ttk, messagebox
import csv, os
from tkinter import filedialog
try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None
from datetime import datetime
try:
    from fpdf import FPDF
except ImportError:
    FPDF = None
try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None 

# =====================================================================================
# 2. CONFIGURACIÓN GLOBAL Y CONSTANTES
# =====================================================================================
FONDO = "#30d647"
FUENTE = ("Arial", 10)
FUENTE_GRANDE = ("Arial", 14) # Fuente más grande para login/registro
rol_global = "Administrador"

# Nombres de archivos CSV
archivo_usuarios = "usuarios.csv"
ventas_archivo = "ventas.csv"
clientes_archivo = "clientes.csv"
proveedores_archivo = "proveedores.csv"
pedidos_archivo = "pedidos.csv"
gastos_archivo = "gastos.csv"
inventario_archivo = "inventario.csv"

# =====================================================================================
# 3. FUNCIONES DE UTILIDAD GENERAL
# =====================================================================================

def crear_ventana_con_fondo(titulo, ancho=1200, alto=800, image_path=None):
    """Crea una nueva ventana Toplevel con una imagen de fondo y devuelve la ventana y el canvas."""
    win = tk.Toplevel()
    win.title(titulo)
    win.geometry(f"{ancho}x{alto}")

    fondo_img = None
    # Intentar cargar la imagen sólo si Pillow está disponible
    if Image is not None and ImageTk is not None:
        try:
            # Usa la ruta de imagen proporcionada o una por defecto si no se especifica.
            if image_path is None:
                default_image_path = r"C:\Users\yorwar\Downloads\tarea de python\fondo de punto de ventas.png"
                image_path = default_image_path

            if os.path.exists(image_path):
                img = Image.open(image_path)
                img = img.resize((ancho, alto), Image.LANCZOS)
                fondo_img = ImageTk.PhotoImage(img)
                win.fondo_img = fondo_img  # Guardar referencia
        except Exception as e:
            fondo_img = None
            print(f"No se pudo cargar la imagen de fondo: {e}")

    canvas = tk.Canvas(win, width=ancho, height=alto)
    canvas.pack(fill="both", expand=True)

    if fondo_img:
        canvas.create_image(0, 0, image=fondo_img, anchor="nw")
    
    return win, canvas

def generar_factura_pdf(factura_id, cliente, articulos, total, vendedor):
    """Genera un archivo PDF para la factura de venta."""
    if FPDF is None:
        messagebox.showerror("Error de Dependencia", 
                             "La librería 'fpdf' no está instalada.\n\n"
                             "Por favor, instálala ejecutando:\npip install fpdf")
        return
    
    try:
        ahora = datetime.now()
        fecha = ahora.strftime("%Y-%m-%d")
        hora = ahora.strftime("%H:%M:%S")

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Encabezado
        pdf.cell(200, 10, txt="FACTURA DE VENTA", ln=True, align="C")
        pdf.ln(5)
        pdf.cell(200, 10, txt="Punto de Venta", ln=True, align="L")
        pdf.cell(200, 10, txt="Sistema de punto de venta", ln=True, align="L")
        pdf.ln(2)
        pdf.cell(200, 10, txt="Dirección: Las flores calle las margaritas edificio la unefa", ln=True, align="L")
        pdf.cell(200, 10, txt="Teléfono: +58 04241796801", ln=True, align="L")
        pdf.cell(200, 10, txt=f"Vendedor: {vendedor}", ln=True, align="L")
        pdf.ln(5)

        # Datos de factura
        pdf.cell(200, 10, txt=f"Número de factura: {factura_id}", ln=True, align="L")
        pdf.cell(200, 10, txt=f"Fecha: {fecha}", ln=True, align="L")
        pdf.cell(200, 10, txt=f"Hora: {hora}", ln=True, align="L")
        pdf.ln(5)
        pdf.cell(200, 10, txt=f"Cliente: {cliente}", ln=True, align="L")
        pdf.ln(5)

        # Tabla de productos
        pdf.set_font("Arial", style="B", size=12)
        pdf.cell(10, 10, "#", 1)
        pdf.cell(60, 10, "Producto", 1)
        pdf.cell(30, 10, "Cantidad", 1)
        pdf.cell(30, 10, "Precio", 1)
        pdf.cell(30, 10, "Total", 1)
        pdf.ln()

        pdf.set_font("Arial", size=12)
        for i, (producto, cantidad, precio, subtotal) in enumerate(articulos, start=1):
            pdf.cell(10, 10, str(i), 1)
            pdf.cell(60, 10, producto, 1)
            pdf.cell(30, 10, str(cantidad), 1)
            pdf.cell(30, 10, f"{precio:.2f}", 1)
            pdf.cell(30, 10, f"{subtotal:.2f}", 1)
            pdf.ln()

        pdf.ln(5)
        pdf.cell(200, 10, txt=f"Total a pagar: {total:.2f} COP", ln=True, align="L")
        pdf.ln(10)

        # Mensaje de cierre
        pdf.cell(200, 10, txt="¡Gracias por tu compra, vuelve pronto!", ln=True, align="L")
        pdf.ln(5)

        # Términos y condiciones
        pdf.set_font("Arial", style="I", size=10)
        pdf.multi_cell(0, 10, txt=(
            "Términos y condiciones:\n"
            "- Los productos comprados no tienen devolución.\n"
            "- Conserve esta factura como su comprobante.\n"
            "- Para más información visite nuestro sitio web."
        ))

        # Definir la carpeta de salida para las facturas
        output_folder = "FacturasGeneradas"
        os.makedirs(output_folder, exist_ok=True)

        # Generar el nombre del archivo y la ruta completa
        file_name = f"factura_{factura_id}_{fecha.replace('-', '')}_{hora.replace(':', '')}.pdf"
        full_path = os.path.join(output_folder, file_name)
        pdf.output(full_path)
        messagebox.showinfo("Factura Generada", f"El archivo '{file_name}' se ha creado exitosamente.")
    except Exception as e:
        messagebox.showerror("Error al Generar Factura", f"No se pudo crear el PDF.\n\nError: {e}")
        messagebox.showerror("Error al Generar Factura", f"No se pudo crear el PDF.\n\nError: {e}")

def actualizar_fecha_hora(label):
    """Actualiza un label con la fecha y hora actual cada segundo."""
    def refrescar():
        ahora = datetime.now()
        fecha = ahora.strftime("📅 %d-%m-%Y")
        hora = ahora.strftime("🕒 %H:%M:%S")
        label.config(text=f"{fecha} {hora}")
        label.after(1000, refrescar)
    refrescar()

def confirmar_cierre(ventana):
    """Muestra un diálogo de confirmación para cerrar sesión."""
    respuesta = messagebox.askyesno("Cerrar sesión", "¿Desea cerrar sesión?")
    if respuesta:
        ventana.destroy()

# =====================================================================================
# 4. GESTIÓN DE MÓDULOS (CRUD)
# =====================================================================================

# -----------------------------------
# MÓDULO DE USUARIOS
# -----------------------------------
def cargar_usuarios():
    if not os.path.exists(archivo_usuarios):
        return []
    with open(archivo_usuarios, newline='', encoding='utf-8') as f:
        return list(csv.reader(f))

def guardar_usuario(usuario):
    with open(archivo_usuarios, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(usuario)

def ventana_agregar(tabla):
    win = tk.Toplevel()
    win.title("Agregar Usuario")
    win.geometry("400x300")
    win.configure(bg=FONDO)
    
    tk.Label(win, text="Usuario", bg=FONDO).pack()
    usuario_entry = tk.Entry(win)
    usuario_entry.pack()

    tk.Label(win, text="Contraseña", bg=FONDO).pack()
    contrasena_entry = tk.Entry(win, show="*")
    contrasena_entry.pack()

    tk.Label(win, text="Rol de usuario", bg=FONDO).pack()
    rol_entry = tk.Entry(win)
    rol_entry.pack()

    def agregar():
        usuario = usuario_entry.get()
        contrasena = contrasena_entry.get()
        rol = rol_entry.get()
        if not (usuario and contrasena and rol):
            messagebox.showwarning("Campos vacíos", "Por favor completa todos los campos.")
            return

        usuarios = cargar_usuarios()
        if any(len(u) >= 2 and u[1] == usuario for u in usuarios):
            messagebox.showwarning("Usuario existente", "El nombre de usuario ya existe.")
        else:
            usuarios = cargar_usuarios()
            nuevo_id = len(usuarios) + 1
            guardar_usuario([nuevo_id, usuario, contrasena, rol])
            tabla.insert("", tk.END, values=[nuevo_id, usuario, "******", rol])
            win.destroy()

    tk.Button(win, text="Guardar", command=agregar).pack(pady=10)

def actualizar_usuario(tabla):
    seleccionado = tabla.selection()
    if not seleccionado:
        messagebox.showwarning("Sin selección", "Selecciona un usuario para actualizar.")
        return

    item = tabla.item(seleccionado)
    valores = item['values'] # Esto es una tupla de los valores de la fila seleccionada
    
    # Asegurarse de que 'valores' tenga suficientes elementos antes de acceder a ellos
    if len(valores) < 4:
        messagebox.showwarning("Datos incompletos", "La fila seleccionada tiene datos incompletos. No se puede actualizar.")
        return
    id_usuario = valores[0]
    usuario_actual = valores[1]
    rol_actual = valores[3]

    win = tk.Toplevel()
    win.title("Actualizar Usuario")
    win.geometry("400x400")
    win.configure(bg=FONDO)

    tk.Label(win, text="Usuario", bg=FONDO).pack()
    usuario_entry = tk.Entry(win)
    usuario_entry.insert(0, usuario_actual)
    usuario_entry.pack()

    tk.Label(win, text="Contraseña (dejar en blanco para no cambiar)", bg=FONDO).pack()
    contrasena_entry = tk.Entry(win, show="*")
    contrasena_entry.pack()

    tk.Label(win, text="Rol de usuario", bg=FONDO).pack()
    rol_entry = tk.Entry(win)
    rol_entry.insert(0, rol_actual)
    rol_entry.pack()

    def guardar_actualizacion():
        nuevo_usuario = usuario_entry.get()
        nueva_contrasena = contrasena_entry.get()
        nuevo_rol = rol_entry.get()

        if not (nuevo_usuario and nuevo_rol):
            messagebox.showwarning("Campos vacíos", "El usuario y el rol no pueden estar vacíos.")
            return

        usuarios = cargar_usuarios()
        for i in range(len(usuarios)):
            if len(usuarios[i]) >= 4 and str(usuarios[i][0]) == str(id_usuario):
                usuarios[i][1] = nuevo_usuario
                if nueva_contrasena:
                    usuarios[i][2] = nueva_contrasena
                usuarios[i][3] = nuevo_rol
                break

        with open(archivo_usuarios, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(usuarios)

        tabla.item(seleccionado, values=[id_usuario, nuevo_usuario, "******", nuevo_rol])
        win.destroy()

    tk.Button(win, text="Guardar Cambios", command=guardar_actualizacion).pack(pady=10)

def eliminar_usuario(tabla):
    seleccionado = tabla.selection()
    if not seleccionado:
        messagebox.showwarning("Sin selección", "Selecciona un usuario para eliminar.")
        return

    respuesta = messagebox.askyesno("Eliminar", "¿Desea eliminar el usuario seleccionado?")
    if not respuesta:
        return

    item = tabla.item(seleccionado)
    valores = item['values']
    id_a_eliminar = str(valores[0])

    usuarios = cargar_usuarios()
    usuarios_filtrados = [u for u in usuarios if u[0] != id_a_eliminar]
    with open(archivo_usuarios, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(usuarios_filtrados)

    tabla.delete(seleccionado)

def abrir_usuarios():
    win, canvas = crear_ventana_con_fondo("ADMINISTRAR USUARIOS")
    ancho_ventana, alto_ventana = 1200, 800

    # Si hay fondo, el color de los elementos será blanco, si no, el color de fondo por defecto.
    if win.fondo_img:
        bg_color_elementos = "#ffffff"
    else:
        canvas.configure(bg=FONDO) # Fallback si no hay imagen
        bg_color_elementos = FONDO

    main_frame = tk.Frame(canvas, bg=bg_color_elementos, bd=2, relief="groove")

    tk.Label(main_frame, text="ADMINISTRAR USUARIOS", font=("Arial", 16, "bold"), bg=bg_color_elementos).pack(pady=10)

    columnas = ("ID", "Usuario", "Contraseña", "Rol de usuario")
    tabla = ttk.Treeview(main_frame, columns=columnas, show="headings")
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=180, anchor="center")
    tabla.pack(pady=20)

    for fila in cargar_usuarios():
        if len(fila) >= 4:
            tabla.insert("", tk.END, values=[fila[0], fila[1], "******", fila[3]])

    frame_botones = tk.Frame(main_frame, bg=bg_color_elementos)
    frame_botones.pack(pady=10)

    tk.Button(frame_botones, text="Agregar Usuario", width=20, command=lambda: ventana_agregar(tabla)).grid(row=0, column=0, padx=10)
    tk.Button(frame_botones, text="Actualizar Usuario", width=20, command=lambda: actualizar_usuario(tabla)).grid(row=0, column=1, padx=10)
    tk.Button(frame_botones, text="Eliminar Usuario", width=20, command=lambda: eliminar_usuario(tabla)).grid(row=0, column=2, padx=10)

    fecha_hora = tk.Label(main_frame, bg=bg_color_elementos, font=FUENTE)
    fecha_hora.pack(pady=10)
    actualizar_fecha_hora(fecha_hora)

    canvas.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame, anchor="center")

# -----------------------------------
# MÓDULO DE VENTAS
# -----------------------------------
# Lista de productos con precios
productos_disponibles = {
    "Reloj Rolex": 5.00,
    "Reloj Casio": 10.00,
    "Reloj Salco": 20.00,
    "Reloj Diesel": 4.00,
    "Cable calibre 4": 2.50,
    "Cable calibre 5": 3.00,
    "Cable calibre 8": 3.50,
    "Adaptador HDMI": 18.00,
    "Pinza multifunción": 35.00,
    "Bolígrafo": 2.50,
    "Mouse inalámbrico": 25.00,
    "Teclado gamer": 45.00,
    "Audífonos Bluetooth": 30.00,
    "Cargador universal": 15.00,
    "Memoria USB 32GB": 12.00,
    "Disco duro 1TB": 55.00,
    "Monitor LED 24''": 120.00,
    "Laptop Lenovo": 450.00,
    "Impresora HP": 80.00,
    "Tablet Samsung": 220.00
}
def cargar_ventas():
    """Carga las ventas desde el archivo ventas.csv."""
    if not os.path.exists(ventas_archivo):
        # Si no existe, crearlo con encabezados
        with open(ventas_archivo, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Factura", "Producto", "Cantidad", "Fecha", "Total", "Costo"])
        return []
    with open(ventas_archivo, newline='', encoding='utf-8') as f:
        return list(csv.reader(f))

def abrir_ventas():
    # Se especifica la ruta de la nueva imagen para esta ventana
    win, canvas = crear_ventana_con_fondo("VENTA DE PRODUCTOS", image_path=r"C:\Users\yorwar\Downloads\tarea de python\imagend de ventas.jpg")
    ancho_ventana, alto_ventana = 1200, 800

    if win.fondo_img:
        bg_color_elementos = "#ffffff"
    else:
        canvas.configure(bg=FONDO)
        bg_color_elementos = FONDO
    main_frame = tk.Frame(canvas, bg=bg_color_elementos, bd=2, relief="groove")

    factura_id = datetime.now().strftime("%Y%m%d%H%M%S")
    total_pagar = tk.DoubleVar(value=0.0)
    articulos = []

    tk.Label(main_frame, text="Cliente:", bg=bg_color_elementos).pack()
    cliente_combo = ttk.Combobox(main_frame, values=["Cliente general"], state="readonly")
    cliente_combo.set("Cliente general")
    cliente_combo.pack()

    tk.Label(main_frame, text="Producto:", bg=bg_color_elementos).pack()
    producto_combo = ttk.Combobox(main_frame, values=list(productos_disponibles.keys()), state="readonly")
    producto_combo.set("Reloj Rolex")
    producto_combo.pack()

    tk.Label(main_frame, text="Cantidad:", bg=bg_color_elementos).pack()
    cantidad_entry = tk.Entry(main_frame)
    cantidad_entry.pack()

    stock_label = tk.Label(main_frame, text="Stock: 100", bg=bg_color_elementos)
    stock_label.pack()

    tabla = ttk.Treeview(main_frame, columns=("Factura", "Cliente", "Producto", "Precio", "Cantidad", "Total"), show="headings")
    for col in ("Factura", "Cliente", "Producto", "Precio", "Cantidad", "Total"):
        tabla.heading(col, text=col)
        tabla.column(col, width=100, anchor="center")
    tabla.pack(pady=10)

    def agregar_articulo():
        try:
            cliente = cliente_combo.get()
            producto = producto_combo.get()
            cantidad = int(cantidad_entry.get())
            precio = productos_disponibles.get(producto, 0)
            total = precio * cantidad
            total_pagar.set(total_pagar.get() + total)
            tabla.insert("", tk.END, values=[factura_id, cliente, producto, f"{precio:.2f}", cantidad, f"{total:.2f}"])
            articulos.append((producto, cantidad, precio, total))
            cantidad_entry.delete(0, tk.END)
        except ValueError:
            messagebox.showerror("Error", "Cantidad inválida")

    tk.Button(main_frame, text="Agregar Artículo", command=agregar_articulo).pack(pady=5)

    tk.Label(main_frame, text=f"Número de Factura: {factura_id}", bg=bg_color_elementos).pack()
    tk.Label(main_frame, text="Precio a Pagar:", bg=bg_color_elementos).pack()
    tk.Label(main_frame, textvariable=total_pagar, bg=bg_color_elementos, font=("Arial", 12, "bold")).pack()

    def realizar_pago():
        if not articulos:
            messagebox.showwarning("Sin artículos", "Agregue al menos un artículo a la venta.")
            return
        
        respuesta = messagebox.askyesno("Confirmar Pago", f"El total a pagar es {total_pagar.get():.2f} bs. ¿Desea continuar?")
        if respuesta:
            generar_factura_pdf(factura_id, cliente_combo.get(), articulos, total_pagar.get(), vendedor=rol_global)
            
            # --- INICIO DE LA CORRECCIÓN --- 
            # Guardar cada artículo vendido en el archivo ventas.csv
            try:
                with open(ventas_archivo, 'a', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    fecha_actual = datetime.now().strftime("%d-%m-%Y")
                    
                    for producto, cantidad, precio, total_articulo in articulos:
                        # Asumimos un costo fijo o lo buscamos en el inventario si es necesario
                        costo_estimado = precio * 0.7 # Ejemplo: costo es 70% del precio
                        writer.writerow([factura_id, producto, cantidad, fecha_actual, total_articulo, costo_estimado * cantidad])
            except Exception as e:
                messagebox.showerror("Error al Guardar Venta", f"No se pudo guardar la venta en el archivo CSV.\n\nError: {e}")
            # --- FIN DE LA CORRECCIÓN ---

            win.destroy()

    tk.Button(main_frame, text="Pagar", command=realizar_pago).pack(pady=10)

    canvas.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame, anchor="center")

# -----------------------------------
# MÓDULO DE CLIENTES
# -----------------------------------
def cargar_clientes():
    if not os.path.exists(clientes_archivo):
        return []
    with open(clientes_archivo, newline='', encoding='utf-8') as f:
        return list(csv.reader(f))

def guardar_cliente(cliente):
    with open(clientes_archivo, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(cliente)

def abrir_clientes():
    win, canvas = crear_ventana_con_fondo("ADMINISTRACIÓN DE CLIENTES")
    ancho_ventana, alto_ventana = 1200, 800

    if win.fondo_img:
        bg_color_elementos = "#ffffff"
    else:
        canvas.configure(bg=FONDO)
        bg_color_elementos = FONDO
    main_frame = tk.Frame(canvas, bg=bg_color_elementos, bd=2, relief="groove")

    tk.Label(main_frame, text="ADMINISTRACIÓN DE CLIENTES", font=("Arial", 16, "bold"), bg=bg_color_elementos).pack(pady=10)

    frame_form = tk.Frame(main_frame, bg=bg_color_elementos)
    frame_form.pack(side="left", padx=20, pady=10, fill="y")

    # Título para la sección de cliente
    tk.Label(frame_form, text="Cliente", font=("Arial", 12, "bold"), bg=bg_color_elementos).pack(pady=(0, 10))

    tk.Label(frame_form, text="Nombre:", bg=bg_color_elementos).pack()
    nombre_entry = tk.Entry(frame_form)
    nombre_entry.pack()

    tk.Label(frame_form, text="Cédula:", bg=bg_color_elementos).pack()
    cedula_entry = tk.Entry(frame_form)
    cedula_entry.pack()

    tk.Label(frame_form, text="Celular:", bg=bg_color_elementos).pack()
    celular_entry = tk.Entry(frame_form)
    celular_entry.pack()

    tk.Label(frame_form, text="Dirección:", bg=bg_color_elementos).pack()
    direccion_entry = tk.Entry(frame_form)
    direccion_entry.pack()

    tk.Label(frame_form, text="Correo:", bg=bg_color_elementos).pack()
    correo_entry = tk.Entry(frame_form)
    correo_entry.pack()

    frame_tabla = tk.Frame(main_frame)
    frame_tabla.pack(side="right", padx=20, pady=10, fill="both", expand=True)

    scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal")
    scroll_y = ttk.Scrollbar(frame_tabla, orient="vertical")

    columnas = ("ID", "Nombre", "Cédula", "Celular", "Dirección", "Correo")
    tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
    
    scroll_x.config(command=tabla.xview)
    scroll_y.config(command=tabla.yview)

    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=120, anchor="center")

    scroll_y.pack(side="right", fill="y")
    scroll_x.pack(side="bottom", fill="x")

    def ingresar_cliente():
        nombre = nombre_entry.get().strip()
        cedula = cedula_entry.get().strip()
        celular = celular_entry.get().strip()
        direccion = direccion_entry.get().strip()
        correo = correo_entry.get().strip()

        if not (nombre and cedula and celular and direccion and correo):
            messagebox.showwarning("Campos vacíos", "Completa todos los campos.")
            return

        clientes = cargar_clientes()
        nuevo_id = len(clientes) + 1
        guardar_cliente([nuevo_id, nombre, cedula, celular, direccion, correo])
        tabla.insert("", tk.END, values=[nuevo_id, nombre, cedula, celular, direccion, correo])

        nombre_entry.delete(0, tk.END)
        cedula_entry.delete(0, tk.END)
        celular_entry.delete(0, tk.END)
        direccion_entry.delete(0, tk.END)
        correo_entry.delete(0, tk.END)

    def eliminar_cliente():
        seleccionado = tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Sin selección", "Selecciona un cliente para eliminar.")
            return

        respuesta = messagebox.askyesno("Eliminar", "¿Desea eliminar el cliente seleccionado?")
        if not respuesta:
            return

        item = tabla.item(seleccionado)
        id_a_eliminar = item['values'][0]

        clientes = cargar_clientes()
        clientes_filtrados = [c for c in clientes if c[0] != str(id_a_eliminar)]

        with open(clientes_archivo, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(clientes_filtrados)

        tabla.delete(seleccionado)

    def modificar_cliente():
        seleccionado = tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Sin selección", "Selecciona un cliente para modificar.")
            return

        item = tabla.item(seleccionado)
        valores = item['values']
        if len(valores) < 6:
            messagebox.showerror("Error", "Cliente incompleto.")
            return

        win = tk.Toplevel()
        win.title("Modificar Cliente")
        win.geometry("400x400")
        win.configure(bg=FONDO)

        tk.Label(win, text="Nombre:", bg=FONDO).pack()
        nombre_mod = tk.Entry(win)
        nombre_mod.insert(0, valores[1])
        nombre_mod.pack()

        tk.Label(win, text="Cédula:", bg=FONDO).pack()
        cedula_mod = tk.Entry(win)
        cedula_mod.insert(0, valores[2])
        cedula_mod.pack()

        tk.Label(win, text="Celular:", bg=FONDO).pack()
        celular_mod = tk.Entry(win)
        celular_mod.insert(0, valores[3])
        celular_mod.pack()

        tk.Label(win, text="Dirección:", bg=FONDO).pack()
        direccion_mod = tk.Entry(win)
        direccion_mod.insert(0, valores[4])
        direccion_mod.pack()

        tk.Label(win, text="Correo:", bg=FONDO).pack()
        correo_mod = tk.Entry(win)
        correo_mod.insert(0, valores[5] if len(valores) > 5 else "") # Añadir correo
        correo_mod.pack()

        def guardar_modificacion():
            nuevo = [
                str(valores[0]),
                nombre_mod.get().strip(),
                cedula_mod.get().strip(),
                celular_mod.get().strip(),
                direccion_mod.get().strip(),
                correo_mod.get().strip()
            ]

            clientes = cargar_clientes()
            for i in range(len(clientes)):
                if clientes[i][0] == str(valores[0]):
                    clientes[i] = nuevo
                    break
                    clientes[i] = nuevo
                    break

            with open(clientes_archivo, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(clientes)

            tabla.item(seleccionado, values=nuevo)
            win.destroy()
        
        tk.Button(win, text="Guardar Cambios", command=guardar_modificacion).pack(pady=10)

    # Cargar datos iniciales en la tabla
    for fila in cargar_clientes():
        if len(fila) >= 6:
            # Asegurarse de que todos los datos se pasen a la tabla
            tabla.insert("", tk.END, values=[fila[0], fila[1], fila[2], fila[3], fila[4], fila[5]])
    tabla.pack(side="left", fill="both", expand=True)

    tk.Button(frame_form, text="Ingresar", command=ingresar_cliente).pack(pady=5)
    tk.Button(frame_form, text="Eliminar", command=eliminar_cliente).pack(pady=5)
    tk.Button(frame_form, text="Modificar", command=modificar_cliente).pack(pady=5)

    canvas.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame, anchor="center")

# -----------------------------------
# MÓDULO DE PROVEEDORES
# -----------------------------------

def cargar_proveedores():
    if not os.path.exists(proveedores_archivo):
        return []
    with open(proveedores_archivo, newline='', encoding='utf-8') as f:
        return list(csv.reader(f))

def guardar_proveedor(proveedor):
    with open(proveedores_archivo, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(proveedor)

def guardar_todos_proveedores(lista):
    with open(proveedores_archivo, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(lista)

def abrir_proveedores():
    # 1. Crear la ventana principal y el canvas de fondo
    win, canvas = crear_ventana_con_fondo("PROVEEDORES")
    ancho_ventana, alto_ventana = 1200, 800

    if win.fondo_img:
        bg_color_elementos = "#ffffff"
    else:
        canvas.configure(bg=FONDO)
        bg_color_elementos = FONDO
    
    # 2. Crear un Frame principal que contendrá todo
    main_frame = tk.Frame(canvas, bg=bg_color_elementos, bd=2, relief="groove")

    tk.Label(main_frame, text="PROVEEDORES", font=("Arial", 16, "bold"), bg=bg_color_elementos).pack(pady=10)
    
    frame_form = tk.Frame(main_frame, bg=bg_color_elementos)
    frame_form.pack(side="left", padx=20, pady=10, fill="y")

    campos = ["Nombre", "Identificación", "Celular", "Dirección", "Correo"]
    entradas = {}

    for campo in campos:
        tk.Label(frame_form, text=campo + ":", bg=bg_color_elementos).pack()
        entrada = tk.Entry(frame_form)
        entrada.pack()
        entradas[campo] = entrada

    # 3. Crear un Frame específico para la tabla y sus scrollbars
    frame_tabla = tk.Frame(main_frame)
    frame_tabla.pack(side="right", padx=20, pady=10, fill="both", expand=True)

    # 4. Crear las barras de desplazamiento DENTRO del frame_tabla
    scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal")
    scroll_y = ttk.Scrollbar(frame_tabla, orient="vertical")

    # 5. Crear la tabla (Treeview) DENTRO del frame_tabla y asociarla a las barras
    columnas = ("Nombre", "Identificación", "Celular", "Dirección", "Correo")
    tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
    
    # 6. Configurar el comando de las barras para que controlen la tabla
    scroll_x.config(command=tabla.xview)
    scroll_y.config(command=tabla.yview)

    # 7. Empaquetar los elementos en el orden correcto: barras a los lados, tabla llena el resto
    #    Este es el paso más crítico.
    scroll_y.pack(side="right", fill="y")
    scroll_x.pack(side="bottom", fill="x")
    tabla.pack(side="left", fill="both", expand=True)
    
    # 8. Configurar las columnas y cargar los datos en la tabla
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=150, anchor="center")

    for fila in cargar_proveedores():
        if len(fila) >= 5:
            tabla.insert("", tk.END, values=fila[:5])

    def registrar():
        datos = [entradas[c].get().strip() for c in campos]
        if any(not d for d in datos):
            messagebox.showwarning("Campos vacíos", "Completa todos los campos.")
            return
        guardar_proveedor(datos)
        tabla.insert("", tk.END, values=datos[:5])
        for entrada in entradas.values():
            entrada.delete(0, tk.END)

    def editar():
        seleccionado = tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Sin selección", "Selecciona un proveedor para editar.")
            return

        item = tabla.item(seleccionado)
        valores = item['values']
        win = tk.Toplevel()
        win.title("Editar Proveedor")
        win.geometry("400x400")
        win.configure(bg=FONDO)

        nuevos = []
        for i, campo in enumerate(columnas):
            tk.Label(win, text=campo + ":", bg=FONDO).pack()
            entrada = tk.Entry(win)
            entrada.insert(0, valores[i])
            entrada.pack()
            nuevos.append(entrada)

        def guardar_edicion():
            actualizado = [e.get().strip() for e in nuevos]
            proveedores = cargar_proveedores()
            for i in range(len(proveedores)):
                if proveedores[i][0] == valores[0]:
                    proveedores[i][:5] = actualizado # Actualizar hasta el correo
                    break
            guardar_todos_proveedores(proveedores)
            tabla.item(seleccionado, values=actualizado)
            win.destroy()

        tk.Button(win, text="Guardar Cambios", command=guardar_edicion).pack(pady=10)

    tk.Button(frame_form, text="Registrar", command=registrar).pack(pady=5)
    tk.Button(frame_form, text="Editar", command=editar).pack(pady=5)

    # 9. Finalmente, colocar el main_frame en el centro del canvas
    canvas.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame, anchor="center")

# -----------------------------------
# MÓDULO DE PEDIDOS
# -----------------------------------
def cargar_pedidos():
    if not os.path.exists(pedidos_archivo):
        return []
    with open(pedidos_archivo, newline='', encoding='utf-8') as f:
        return list(csv.reader(f))
        return list(csv.reader(f))

def guardar_pedido(pedido):
    with open(pedidos_archivo, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(pedido)
        writer = csv.writer(f)
        writer.writerow(pedido)

def actualizar_stock(producto, cantidad):
    inventario = cargar_inventario()
    for i in range(len(inventario)):
        if len(inventario[i]) >= 6 and inventario[i][1] == producto:
            try:
                inventario[i][5] = str(int(inventario[i][5]) + cantidad)
            except ValueError:
                continue
    with open(inventario_archivo, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(inventario)
        writer = csv.writer(f)
        writer.writerows(inventario)

def abrir_pedidos():
    win, canvas = crear_ventana_con_fondo("PEDIDOS")
    ancho_ventana, alto_ventana = 1200, 800

    if win.fondo_img:
        bg_color_elementos = "#ffffff"
    else:
        canvas.configure(bg=FONDO)
        bg_color_elementos = FONDO
    main_frame = tk.Frame(canvas, bg=bg_color_elementos, bd=2, relief="groove")

    tk.Label(main_frame, text="PEDIDOS", font=("Arial", 16, "bold"), bg=bg_color_elementos).pack(pady=10)

    frame_form = tk.Frame(main_frame, bg=bg_color_elementos)
    frame_form.pack(side="left", padx=20, pady=10, fill="y")

    tk.Label(frame_form, text="N° Pedido:", bg=bg_color_elementos).pack()
    pedido_entry = tk.Entry(frame_form)
    pedido_entry.insert(0, "")
    pedido_entry.pack()

    tk.Label(frame_form, text="Proveedor:", bg=bg_color_elementos).pack()
    proveedores = [p[0] for p in cargar_proveedores()]
    proveedor_combo = ttk.Combobox(frame_form, values=proveedores, state="readonly")
    proveedor_combo.set(proveedores[0] if proveedores else "")
    proveedor_combo.pack()

    tk.Label(frame_form, text="Producto:", bg=bg_color_elementos).pack()
    productos = [p[1] for p in cargar_inventario() if len(p) >= 6]
    producto_combo = ttk.Combobox(frame_form, values=productos, state="readonly")
    producto_combo.set(productos[0] if productos else "")
    producto_combo.pack()

    tk.Label(frame_form, text="Nueva Cant.:", bg=bg_color_elementos).pack()
    cantidad_entry = tk.Entry(frame_form)
    cantidad_entry.pack()

    tabla = ttk.Treeview(main_frame, columns=("Pedido", "Proveedor", "Producto", "Cantidad", "Fecha", "Hora"), show="headings")
    for col in ("Pedido", "Proveedor", "Producto", "Cantidad", "Fecha", "Hora"):
        tabla.heading(col, text=col)
        tabla.column(col, width=120, anchor="center")
    tabla.pack(side="right", padx=20, pady=10, fill="both", expand=True)

    def agregar():
        try:
            pedido = int(pedido_entry.get())
            proveedor = proveedor_combo.get()
            producto = producto_combo.get()
            cantidad = int(cantidad_entry.get())
        except:
            messagebox.showerror("Error", "Datos inválidos.")
            return
        tabla.insert("", tk.END, values=(pedido, proveedor, producto, cantidad, "", ""))
        cantidad_entry.delete(0, tk.END)

    def registrar():
        try:
            pedido = int(pedido_entry.get())
            proveedor = proveedor_combo.get()
            producto = producto_combo.get()
            cantidad = int(cantidad_entry.get())
        except:
            messagebox.showerror("Error", "Datos inválidos.")
            return

        fecha = datetime.now().strftime("%d-%m-%Y")
        hora = datetime.now().strftime("%H:%M:%S")
        guardar_pedido([pedido, proveedor, producto, cantidad, fecha, hora])
        actualizar_stock(producto, cantidad)

        messagebox.showinfo("Pedido registrado", "El pedido ha sido registrado exitosamente.")

        pedido_entry.delete(0, tk.END)
        proveedor_combo.set("")
        producto_combo.set("")
        cantidad_entry.delete(0, tk.END)

    def ver():
        tabla.delete(*tabla.get_children())
        for fila in cargar_pedidos():
            if len(fila) >= 6:
                tabla.insert("", tk.END, values=fila)

    tk.Button(frame_form, text="Agregar", command=agregar).pack(pady=5)
    tk.Button(frame_form, text="Registrar", command=registrar).pack(pady=5)
    tk.Button(frame_form, text="Ver pedidos", command=ver).pack(pady=5)

    canvas.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame, anchor="center")

# -----------------------------------
# MÓDULO DE REPORTES
# -----------------------------------
def abrir_reportes():
    win, canvas = crear_ventana_con_fondo("REPORTES", ancho=1300) # Aumentar ancho para más espacio
    ancho_ventana, alto_ventana = 1200, 800

    bg_color_elementos = "#ffffff" if win.fondo_img else FONDO # Color de fondo para widgets

    # Cargar la imagen aquí, después de crear la ventana Toplevel
    try:
        reportes_logo = tk.PhotoImage(file="icono_calendario.png") # Asegúrate que este archivo exista
        win.reportes_logo = reportes_logo # Guardar referencia para evitar que se elimine
    except tk.TclError:
        reportes_logo = None # La imagen no se pudo cargar, los botones no la mostrarán

    main_frame = tk.Frame(canvas, bg=bg_color_elementos, bd=2, relief="groove")

    tk.Label(main_frame, text="📊 REPORTES", font=("Arial", 18, "bold"), bg=bg_color_elementos).pack(pady=10)

    # Frame para ambos reportes
    reportes_container = tk.Frame(main_frame, bg=bg_color_elementos)
    reportes_container.pack(pady=20, padx=20, fill="both", expand=True)

    # --- Reporte de Ventas ---
    frame_ventas = tk.LabelFrame(reportes_container, text="Reporte de Ventas Totales", font=("Arial", 12, "bold"), bg=bg_color_elementos, padx=10, pady=10)
    frame_ventas.pack(side="left", padx=20, fill="both", expand=True)

    tk.Label(frame_ventas, text="Desde:", bg=bg_color_elementos).grid(row=1, column=0, sticky="e")
    desde_ventas = tk.Entry(frame_ventas)
    desde_ventas.insert(0, datetime.now().strftime("%Y-%m-%d"))
    desde_ventas.grid(row=1, column=1)

    tk.Label(frame_ventas, text="Hasta:", bg=bg_color_elementos).grid(row=2, column=0, sticky="e")
    hasta_ventas = tk.Entry(frame_ventas)
    hasta_ventas.insert(0, datetime.now().strftime("%Y-%m-%d"))
    hasta_ventas.grid(row=2, column=1)

    def filtrar_ventas():
        # 1. Limpiar la tabla antes de mostrar nuevos resultados
        for item in tabla_ventas.get_children():
            tabla_ventas.delete(item)
        
        # 2. Obtener y validar las fechas del formulario
        try:
            desde = datetime.strptime(desde_ventas.get(), "%Y-%m-%d")
            hasta = datetime.strptime(hasta_ventas.get(), "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error de formato", "La fecha debe estar en formato YYYY-MM-DD.")
            return

        # 3. Cargar los datos de ventas y procesarlos
        ventas_registradas = cargar_ventas()
        total_ventas_periodo = 0
        for venta in ventas_registradas:
            # Asegurarse de que la fila tiene la estructura correcta (al menos 5 columnas y una fecha)
            if len(venta) >= 5 and venta[3]:
                try:
                    # Corregido: Leer la fecha en formato DD-MM-YYYY del CSV
                    fecha_venta = datetime.strptime(venta[3], "%d-%m-%Y")
                    if desde <= fecha_venta <= hasta:
                        # 4. Insertar cada producto vendido en la tabla
                        producto, cantidad, total = venta[1], venta[2], float(venta[4])
                        tabla_ventas.insert("", tk.END, values=(producto, cantidad, f"{total:,.2f}"))
                        total_ventas_periodo += total
                except (ValueError, IndexError):
                    continue # Ignorar filas con formato de fecha o número incorrecto
        
        total_ventas_label.config(text=f"Total General de Ventas: {total_ventas_periodo:,.2f} COP")

    # Tabla de ventas modificada para mostrar detalles
    tabla_ventas = ttk.Treeview(frame_ventas, columns=("Producto", "Cantidad", "Total"), show="headings")
    tabla_ventas.heading("Producto", text="Producto Vendido")
    tabla_ventas.heading("Cantidad", text="Cantidad")
    tabla_ventas.heading("Total", text="Total Venta")
    tabla_ventas.column("Producto", width=200, anchor="w")
    tabla_ventas.column("Cantidad", width=80, anchor="center")
    tabla_ventas.column("Total", width=120, anchor="e")
    tabla_ventas.grid(row=4, columnspan=2, pady=10, sticky="nsew")

    tk.Button(frame_ventas, text="Filtrar", image=reportes_logo, compound="left", command=filtrar_ventas).grid(row=3, columnspan=2, pady=5)
    total_ventas_label = tk.Label(frame_ventas, text="Total General de Ventas: 0.00 COP", font=("Arial", 11, "bold"), bg=bg_color_elementos)
    total_ventas_label.grid(row=5, columnspan=2, pady=(10,0))
    tk.Label(frame_ventas, text="Muestra los productos vendidos en el período seleccionado.", bg=bg_color_elementos, wraplength=400).grid(row=6, columnspan=2)

    # --- Reporte de Ganancias ---
    frame_ganancias = tk.LabelFrame(reportes_container, text="Reporte de Ganancias", font=("Arial", 12, "bold"), bg=bg_color_elementos, padx=10, pady=10)
    frame_ganancias.pack(side="right", padx=20, fill="both", expand=True)

    tk.Label(frame_ganancias, text="Desde:", bg=bg_color_elementos).grid(row=1, column=0, sticky="e") # Corregido
    desde_ganancias = tk.Entry(frame_ganancias)
    desde_ganancias.insert(0, datetime.now().strftime("%Y-%m-%d"))
    desde_ganancias.grid(row=1, column=1)

    tk.Label(frame_ganancias, text="Hasta:", bg=bg_color_elementos).grid(row=2, column=0, sticky="e") # Corregido
    hasta_ganancias = tk.Entry(frame_ganancias)
    hasta_ganancias.insert(0, datetime.now().strftime("%Y-%m-%d"))
    hasta_ganancias.grid(row=2, column=1)

    def calcular_ganancias():
        for item in tabla_ganancias.get_children():
            tabla_ganancias.delete(item)

        try:
            desde = datetime.strptime(desde_ganancias.get(), "%Y-%m-%d")
            hasta = datetime.strptime(hasta_ganancias.get(), "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error de formato", "La fecha debe estar en formato YYYY-MM-DD.")
            return

        ventas = cargar_ventas()
        ganancias_por_dia = {}
        ganancia_total_periodo = 0
        for v in ventas:
            if len(v) >= 6 and v[3]:
                try:
                    # Corregido: Leer la fecha en formato DD-MM-YYYY del CSV
                    fecha_venta = datetime.strptime(v[3], "%d-%m-%Y")
                except ValueError:
                    continue

                if desde <= fecha_venta <= hasta:
                    try:
                        ganancia_venta = float(v[4]) - float(v[5])
                        fecha_str = fecha_venta.strftime("%Y-%m-%d")
                        ganancias_por_dia[fecha_str] = ganancias_por_dia.get(fecha_str, 0) + ganancia_venta
                        ganancia_total_periodo += ganancia_venta
                    except (ValueError, IndexError):
                        continue
        
        for fecha, ganancia_dia in sorted(ganancias_por_dia.items()):
            tabla_ganancias.insert("", tk.END, values=(fecha, f"{ganancia_dia:,.2f}"))
        
        # Añadir fila de total
        tabla_ganancias.insert("", tk.END, values=("GANANCIA TOTAL", f"{ganancia_total_periodo:,.2f}"), tags=('total_row',))

    tabla_ganancias = ttk.Treeview(frame_ganancias, columns=("Fecha", "Ganancia"), show="headings")
    tabla_ganancias.heading("Fecha", text="Fecha")
    tabla_ganancias.heading("Ganancia", text="Ganancia del Día (COP)")
    tabla_ganancias.column("Fecha", width=150, anchor="center")
    tabla_ganancias.column("Ganancia", width=200, anchor="e")
    tabla_ganancias.grid(row=4, columnspan=2, pady=10, sticky="nsew")
    tabla_ganancias.tag_configure('total_row', font=('Arial', 10, 'bold'))

    tk.Button(frame_ganancias, text="Reporte", image=reportes_logo, compound="left", command=calcular_ganancias).grid(row=3, columnspan=2, pady=5)
    tk.Label(frame_ganancias, text="Muestra la ganancia (venta - costo) por día y el total del período.", bg=bg_color_elementos, wraplength=400).grid(row=5, columnspan=2)

    canvas.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame, anchor="center")
# -----------------------------------
# MÓDULO DE GASTOS
# -----------------------------------
def cargar_gastos():
    if not os.path.exists(gastos_archivo):
        return []
    with open(gastos_archivo, newline='', encoding='utf-8') as f:
        return list(csv.reader(f))
        return list(csv.reader(f))

def guardar_gasto(gasto):
    with open(gastos_archivo, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(gasto)
        writer = csv.writer(f)
        writer.writerow(gasto)

def abrir_gastos():
    win, canvas = crear_ventana_con_fondo("CONTROL DE GASTOS")
    ancho_ventana, alto_ventana = 1200, 800

    if win.fondo_img:
        bg_color_elementos = "#ffffff"
    else:
        canvas.configure(bg=FONDO)
        bg_color_elementos = FONDO
    main_frame = tk.Frame(canvas, bg=bg_color_elementos, bd=2, relief="groove")

    tk.Label(main_frame, text="CONTROL DE GASTOS", font=("Arial", 16, "bold"), bg=bg_color_elementos).pack(pady=10)

    frame_form = tk.Frame(main_frame, bg=bg_color_elementos)
    frame_form.pack(side="left", padx=20, pady=10, fill="y")

    tk.Label(frame_form, text="Concepto:", bg=bg_color_elementos).pack()
    concepto_entry = tk.Entry(frame_form)
    concepto_entry.pack()

    tk.Label(frame_form, text="Valor:", bg=bg_color_elementos).pack()
    valor_entry = tk.Entry(frame_form)
    valor_entry.pack()

    tk.Label(frame_form, text="Entidad:", bg=bg_color_elementos).pack()
    entidad_entry = tk.Entry(frame_form)
    entidad_entry.pack()

    tk.Label(frame_form, text="Fecha:", bg=bg_color_elementos).pack()
    fecha_entry = tk.Entry(frame_form)
    fecha_entry.insert(0, datetime.now().strftime("%d-%m-%Y"))
    fecha_entry.pack()

    columnas = ("ID", "Concepto", "Valor", "Entidad", "Fecha")
    tabla = ttk.Treeview(main_frame, columns=columnas, show="headings")
    
    tabla.heading("ID", text="ID")
    tabla.column("ID", width=50, anchor="center")
    tabla.heading("Concepto", text="Concepto")
    tabla.column("Concepto", width=250, anchor="w")
    tabla.heading("Valor", text="Valor")
    tabla.column("Valor", width=120, anchor="e")
    tabla.heading("Entidad", text="Entidad")
    tabla.column("Entidad", width=200, anchor="w")
    tabla.heading("Fecha", text="Fecha")
    tabla.column("Fecha", width=120, anchor="center")
    tabla.pack(side="right", padx=20, pady=10, fill="both", expand=True)

    for fila in cargar_gastos():
        if len(fila) == 5:
            tabla.insert("", tk.END, values=fila)

    def ingresar_gasto():
        concepto = concepto_entry.get().strip()
        valor = valor_entry.get().strip()
        entidad = entidad_entry.get().strip()
        fecha = fecha_entry.get().strip()

        if not (concepto and valor and entidad and fecha):
            messagebox.showwarning("Campos vacíos", "Completa todos los campos.")
            return

        try:
            valor_float = float(valor)
        except ValueError:
            messagebox.showerror("Error", "El valor debe ser numérico.")
            return

        gastos = cargar_gastos()
        nuevo_id = len(gastos) + 1
        nuevo_gasto = [nuevo_id, concepto, valor_float, entidad, fecha]
        guardar_gasto(nuevo_gasto)
        tabla.insert("", tk.END, values=nuevo_gasto)

        messagebox.showinfo("Éxito", f"Gasto registrado correctamente. Valor: {valor_float:,.0f} COP")

        concepto_entry.delete(0, tk.END)
        valor_entry.delete(0, tk.END)
        entidad_entry.delete(0, tk.END)
        fecha_entry.delete(0, tk.END)
        fecha_entry.insert(0, datetime.now().strftime("%d-%m-%Y"))

    def modificar_gasto():
        seleccionado = tabla.selection()
        if not seleccionado:
            messagebox.showwarning("Sin selección", "Selecciona un gasto para modificar.")
            return

        item = tabla.item(seleccionado)
        valores = item['values']
        win = tk.Toplevel()
        win.title("Modificar Gasto")
        win.geometry("400x400")
        win.configure(bg=FONDO)

        campos = ["Concepto", "Valor", "Entidad", "Fecha"]
        entradas = []

        for i, campo in enumerate(campos):
            tk.Label(win, text=campo + ":", bg=FONDO).pack()
            entrada = tk.Entry(win)
            entrada.insert(0, valores[i + 1])
            entrada.pack()
            entradas.append(entrada)

        def guardar_modificacion():
            try:
                nuevo_valor = float(entradas[1].get().strip())
            except ValueError:
                messagebox.showerror("Error", "El valor debe ser numérico.")
                return

            nuevo = [
                valores[0],
                entradas[0].get().strip(),
                nuevo_valor,
                entradas[2].get().strip(),
                entradas[3].get().strip()
            ]

            gastos = cargar_gastos()
            for i in range(len(gastos)):
                if str(gastos[i][0]) == str(valores[0]):
                    gastos[i] = nuevo
                    break

            with open(gastos_archivo, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(gastos)

            tabla.item(seleccionado, values=nuevo)
            win.destroy()

        tk.Button(win, text="Guardar Cambios", command=guardar_modificacion).pack(pady=10)

    tk.Button(frame_form, text="Ingresar", command=ingresar_gasto).pack(pady=5)
    tk.Button(frame_form, text="Modificar", command=modificar_gasto).pack(pady=5)

    canvas.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame, anchor="center")
def cargar_inventario():
    if not os.path.exists(inventario_archivo):
        # Si no existe, crearlo con encabezados
        with open(inventario_archivo, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Id", "Producto", "Proveedor", "Precio", "Costo", "Stock"])
        return []
    with open(inventario_archivo, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None) # Omitir encabezado
        return list(reader)

def guardar_inventario(productos):
    with open(inventario_archivo, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Id", "Producto", "Proveedor", "Precio", "Costo", "Stock"]) # Escribir encabezado
        writer.writerows(productos)

# -----------------------------------
# MÓDULO DE INVENTARIO
# -----------------------------------
def abrir_inventario():
    win, canvas = crear_ventana_con_fondo("INVENTARIOS")
    ancho_ventana, alto_ventana = 1200, 800

    if win.fondo_img:
        bg_color_elementos = "#ffffff"
    else:
        canvas.configure(bg=FONDO)
        bg_color_elementos = FONDO

    # --- Funciones internas para evitar variables globales ---
    def actualizar_total():
        total = 0
        for item_id in tabla.get_children():
            try:
                valores = tabla.item(item_id)['values']
                if len(valores) >= 6:
                    # Limpiar el string de precio ('$20.00' -> '20.00') antes de convertir
                    precio_str = str(valores[3]).replace('$', '').strip()
                    precio = float(precio_str)
                    stock = int(valores[5])
                    total += precio * stock
            except (ValueError, IndexError):
                continue
        total_label.config(text=f"Total en Inventario: ${total:,.2f}")

    def ingresar_producto():
        nombre = nombre_entry.get().strip()
        proveedor = proveedor_combo.get()
        # Limpiar el símbolo '$' si el usuario lo ingresa
        precio_str = precio_entry.get().strip().replace('$', '')
        costo_str = costo_entry.get().strip().replace('$', '')
        stock_str = stock_entry.get().strip()

        if not all((nombre, proveedor, precio_str, costo_str, stock_str)):
            messagebox.showwarning("Campos vacíos", "Completa todos los campos.")
            return

        try:
            precio = float(precio_str)
            costo = float(costo_str)
            stock = int(stock_str)
        except ValueError:
            messagebox.showerror("Error de Formato", "Precio, costo y stock deben ser números válidos.")
            return

        productos = cargar_inventario()
        nuevo_id = max([int(p[0]) for p in productos if p and p[0].isdigit()] + [0]) + 1 # Generar nuevo ID
        
        # Guardar en CSV sin '$', pero mostrar en tabla con '$'
        producto_csv = [nuevo_id, nombre, proveedor, f"{precio:.2f}", f"{costo:.2f}", stock]
        producto_tabla = [nuevo_id, nombre, proveedor, f"${precio:.2f}", f"${costo:.2f}", stock]
        
        productos.append(producto_csv)
        guardar_inventario(productos)
        tabla.insert("", tk.END, values=producto_tabla)
        actualizar_total()

        # Limpiar campos
        nombre_entry.delete(0, tk.END)
        precio_entry.delete(0, tk.END)
        costo_entry.delete(0, tk.END)
        stock_entry.delete(0, tk.END)

    def editar_producto(tabla_ref):
        seleccionado = tabla_ref.selection()
        if not seleccionado:
            messagebox.showwarning("Sin selección", "Selecciona un producto para editar.")
            return

        item = tabla_ref.item(seleccionado)
        valores = item['values']
        if len(valores) < 6:
            messagebox.showerror("Error de Datos", "El producto seleccionado tiene datos incompletos.")
            return

        win_edit = tk.Toplevel()
        win_edit.title("Editar Producto")
        win_edit.geometry("800x800")
        win_edit.configure(bg=FONDO)

        tk.Label(win_edit, text="Nombre:", bg=FONDO).pack() # No change here, just context
        nombre_edit_entry = tk.Entry(win_edit)
        nombre_edit_entry.insert(0, valores[1])
        nombre_edit_entry.pack()

        tk.Label(win_edit, text="Proveedor:", bg=FONDO).pack()
        proveedor_edit_combo = ttk.Combobox(win_edit, values=["Proveedor1", "Proveedor2", "Proveedor3"], state="readonly")
        proveedor_edit_combo.set(valores[2])
        proveedor_edit_combo.pack()

        tk.Label(win_edit, text="Precio ($):", bg=FONDO).pack()
        precio_edit_entry = tk.Entry(win_edit)
        precio_edit_entry.insert(0, valores[3])
        precio_edit_entry.pack()

        tk.Label(win_edit, text="Costo ($):", bg=FONDO).pack()
        costo_edit_entry = tk.Entry(win_edit)
        costo_edit_entry.insert(0, valores[4])
        costo_edit_entry.pack()

        tk.Label(win_edit, text="Stock:", bg=FONDO).pack() # No change here, just context
        stock_edit_entry = tk.Entry(win_edit)
        stock_edit_entry.insert(0, valores[5])
        stock_edit_entry.pack()

        def guardar_edicion():
            try:
                # Limpiar el símbolo '$' al guardar la edición
                nuevos_valores = [
                    valores[0], # Mantener el ID original
                    nombre_edit_entry.get().strip(),
                    proveedor_edit_combo.get(),
                    # Guardar en CSV sin '$'
                    f"{float(precio_edit_entry.get().strip().replace('$', '')):.2f}",
                    f"{float(costo_edit_entry.get().strip().replace('$', '')):.2f}",
                    int(stock_edit_entry.get())
                ]
            except ValueError:
                messagebox.showerror("Error de Formato", "Precio, costo y stock deben ser números válidos.")
                return

            productos = cargar_inventario()
            for i, p in enumerate(productos):
                if str(p[0]) == str(valores[0]):
                    productos[i] = nuevos_valores
                    break
            
            guardar_inventario(productos)
            # Actualizar la tabla con el formato '$'
            valores_tabla = nuevos_valores.copy()
            valores_tabla[3] = f"${float(valores_tabla[3]):.2f}"
            valores_tabla[4] = f"${float(valores_tabla[4]):.2f}"
            
            tabla_ref.item(seleccionado, values=valores_tabla)
            actualizar_total()
            win_edit.destroy()

        tk.Button(win_edit, text="Guardar Cambios", command=guardar_edicion).pack(pady=10)

    def eliminar_producto(tabla_ref):
        seleccionado = tabla_ref.selection()
        if not seleccionado:
            messagebox.showwarning("Sin selección", "Selecciona un producto para eliminar.")
            return

        if messagebox.askyesno("Confirmar", "¿Estás seguro de que quieres eliminar este producto?"):
            id_a_eliminar = tabla_ref.item(seleccionado)['values'][0]
            productos = [p for p in cargar_inventario() if str(p[0]) != str(id_a_eliminar)]
            guardar_inventario(productos)
            tabla_ref.delete(seleccionado)
            actualizar_total()

    def refrescar_inventario_completo():
        """Borra la tabla y la vuelve a cargar desde el archivo CSV."""
        # Borrar todos los items actuales
        for item in tabla.get_children():
            tabla.delete(item)
        # Cargar y mostrar los datos actualizados
        for fila in cargar_inventario():
            if len(fila) == 6:
                fila_formateada = fila.copy()
                fila_formateada[3] = f"${float(fila[3]):.2f}" # Añadir $ al precio
                fila_formateada[4] = f"${float(fila[4]):.2f}" # Añadir $ al costo
                tabla.insert("", tk.END, values=fila_formateada)
        actualizar_total()
        messagebox.showinfo("Inventario Actualizado", "La lista de productos ha sido actualizada.")

    def exportar_inventario_excel():
        """Exporta los datos del inventario a un archivo de Excel (.xlsx)."""
        if Workbook is None:
            messagebox.showerror("Dependencia Faltante", 
                                 "La librería 'openpyxl' es necesaria para exportar a Excel.\n\n"
                                 "Por favor, instálala ejecutando:\npip install openpyxl")
            return

        inventario = cargar_inventario()
        if not inventario:
            messagebox.showwarning("Sin Datos", "No hay datos en el inventario para exportar.")
            return

        try:
            # Pedir al usuario dónde guardar el archivo
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
                title="Guardar inventario como...",
                initialfile="reporte_inventario.xlsx"
            )

            if not filepath: # Si el usuario cancela
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"

            # Escribir encabezados
            headers = ["Id", "Producto", "Proveedor", "Precio", "Costo", "Stock"]
            ws.append(headers)

            # Escribir datos del inventario
            for producto in inventario:
                ws.append(producto)
            
            wb.save(filepath)
            messagebox.showinfo("Exportación Exitosa", f"El inventario ha sido exportado a:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"Ocurrió un error al exportar a Excel:\n{e}")

    # --- Interfaz de la ventana de inventario ---
    main_frame = tk.Frame(canvas, bg=bg_color_elementos, bd=2, relief="groove")

    tk.Label(main_frame, text="INVENTARIOS", font=("Arial", 16, "bold"), bg=bg_color_elementos).pack(pady=10)

    frame_form = tk.Frame(main_frame, bg=bg_color_elementos)
    frame_form.pack(side="left", padx=20, pady=10, fill="y")

    tk.Label(frame_form, text="Producto", font=("Arial", 12, "bold"), bg=bg_color_elementos).pack(pady=(0, 10))

    tk.Label(frame_form, text="Nombre:", bg=bg_color_elementos).pack()
    nombre_entry = tk.Entry(frame_form)
    nombre_entry.pack()

    tk.Label(frame_form, text="Proveedor:", bg=bg_color_elementos).pack()
    proveedor_combo = ttk.Combobox(frame_form, values=["Proveedor1", "Proveedor2", "Proveedor3"], state="readonly")
    proveedor_combo.set("Proveedor1")
    proveedor_combo.pack()

    tk.Label(frame_form, text="Precio ($):", bg=bg_color_elementos).pack()
    precio_entry = tk.Entry(frame_form)
    precio_entry.pack()

    tk.Label(frame_form, text="Costo ($):", bg=bg_color_elementos).pack()
    costo_entry = tk.Entry(frame_form)
    costo_entry.pack()

    tk.Label(frame_form, text="Stock:", bg=bg_color_elementos).pack() # No change here, just context
    stock_entry = tk.Entry(frame_form)
    stock_entry.pack()

    tk.Button(frame_form, text="➕ Ingresar Producto", command=ingresar_producto, width=20).pack(pady=10)
    tk.Button(frame_form, text="✏️ Editar Producto", command=lambda: editar_producto(tabla), width=20).pack(pady=5)
    tk.Button(frame_form, text="🗑️ Eliminar Producto", command=lambda: eliminar_producto(tabla), width=20).pack(pady=5)
    tk.Button(frame_form, text="🔄 Actualizar Inventario", command=refrescar_inventario_completo, width=20).pack(pady=5)
    tk.Button(frame_form, text="📄 Exportar Inventario", command=exportar_inventario_excel, width=20).pack(pady=5)

    frame_tabla = tk.Frame(main_frame, bg=bg_color_elementos)
    frame_tabla.pack(side="right", padx=20, pady=10, expand=True, fill="both")

    columnas = ("Id", "Producto", "Proveedor", "Precio ($)", "Costo ($)", "Stock")
    tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings")
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=100, anchor="center")
    tabla.pack(expand=True, fill="both")

    # Cargar datos iniciales con formato de dólar
    for fila in cargar_inventario():
        if len(fila) == 6:
            fila[3] = f"${float(fila[3]):.2f}"
            fila[4] = f"${float(fila[4]):.2f}"
            tabla.insert("", tk.END, values=fila)

    total_label = tk.Label(frame_tabla, text="Total en Inventario: $0.00", bg=bg_color_elementos, font=("Arial", 12, "bold"))
    total_label.pack(pady=10)

    canvas.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame, anchor="center")

# =====================================================================================
# 5. VENTANAS PRINCIPALES DE LA APLICACIÓN
# =====================================================================================

def abrir_about_us():
    ventana = tk.Toplevel()
    ventana.title("InnovaciónCode9889")
    ventana.geometry("800x600")
    ventana.configure(bg=FONDO)
    
    tk.Label(ventana, text="INNOVASIÓNCODE", font=("Arial", 18, "bold"), bg=FONDO).pack(pady=10)

    logo = tk.Label(ventana, text="INNOVASIÓN<>CODE", font=("Arial", 16, "bold"), bg=FONDO)
    logo.pack(pady=5)

    descripcion = (
        "Somos  un equipo de trabajo innovando y buscando  soluciones efetiva  comprometida para  brindar soluciones tecnológicas innovadoras.\n"
        "Nuestro equipo está dedicado a proporcionar productos y servicios que favorecen  a los cliente.\n"
        "y con dedicación y trabajo logramos lo que se queria ."
    )
    tk.Label(ventana, text=descripcion, font=("Arial", 11), bg=FONDO, justify="center").pack(pady=10)

    info = (
        "Prototipo  final   Sistema  Punto de Venta\n"
        "Versión: 3.4.0\n"
        "Última actualización: 25/11/2025"
    )
    tk.Label(ventana, text=info, font=("Arial", 11), bg=FONDO, justify="center").pack(pady=5)

    soporte = (
        "Soporte: @yorwar , @mauricio , @yoan.com\n"
        "Celular: +58 04262381791\n"
        "Innovasión Code"
    )
    tk.Label(ventana, text=soporte, font=("Arial", 11, "bold"), bg=FONDO, justify="center").pack(pady=5)

    copyright = "Copyright © 2026 Todos los derechos reservados"
    tk.Label(ventana, text=copyright, font=("Arial", 10), bg=FONDO).pack(pady=10)

    fecha_hora = tk.Label(ventana, bg=FONDO, font=FUENTE)
    fecha_hora.place(relx=1.0, rely=0, x=-10, y=10, anchor="ne")
    actualizar_fecha_hora(fecha_hora)

def abrir_punto_de_venta():
    ventana_pdv, canvas_pdv = crear_ventana_con_fondo("PUNTO DE VENTA VERSION 3.1.0")
    ancho_ventana, alto_ventana = 1200, 800

    bg_color_elementos = "#ffffff" if ventana_pdv.fondo_img else FONDO

    main_frame_pdv = tk.Frame(canvas_pdv, bg=bg_color_elementos, bd=2, relief="groove")

    # Botón de cerrar sesión en la esquina superior derecha de la ventana
    btn_cerrar = tk.Button(canvas_pdv, text="🔐 Cerrar Sesión", font=("Arial", 10, "bold"),
                           command=lambda: confirmar_cierre(ventana_pdv),
                           bg="lightgray", relief="raised")
    btn_cerrar.place(relx=1.0, rely=0, x=-10, y=10, anchor="ne")

    botones = [
        "Ventas", "Inventario", "Clientes", "Proveedor", "Pedidos",
        "Reportes", "Gastos", "Usuarios", "About Us" 
    ]

    grid_frame = tk.Frame(main_frame_pdv, bg=bg_color_elementos)
    grid_frame.pack(pady=(40, 20), padx=20)

    for i, texto in enumerate(botones):
        cmd = None
        if texto == "Usuarios": cmd = abrir_usuarios 
        elif texto == "Ventas": cmd = abrir_ventas 
        elif texto == "Inventario": cmd = abrir_inventario 
        elif texto == "Clientes": cmd = abrir_clientes 
        elif texto == "Proveedor": cmd = abrir_proveedores 
        elif texto == "Pedidos": cmd = abrir_pedidos 
        elif texto == "Reportes": cmd = abrir_reportes 
        elif texto == "Gastos": cmd = abrir_gastos 
        elif texto == "About Us": cmd = abrir_about_us 

        btn = tk.Button(grid_frame, text=texto, width=15, command=cmd, font=("Arial", 11, "bold"))
        btn.grid(row=i // 2, column=i % 2, padx=10, pady=5)

    tk.Label(main_frame_pdv, text=f"Bienvenido: {rol_global}", bg=bg_color_elementos, font=("Arial", 12, "bold")).pack(pady=10)
    
    fecha_hora = tk.Label(main_frame_pdv, bg=bg_color_elementos, font=FUENTE)
    fecha_hora.pack(side="bottom", pady=10)
    actualizar_fecha_hora(fecha_hora)

    canvas_pdv.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame_pdv, anchor="center")

def abrir_login():
    # Se pasa la ruta de la imagen para que la ventana de inicio de sesión tenga el mismo fondo
    ventana_login, canvas_login = crear_ventana_con_fondo("Iniciar Sesión", ancho=800, alto=600, image_path=r"C:\Users\yorwar\Downloads\tarea de python\imagen de aviones.jpg")
    ancho_ventana, alto_ventana = 800, 600

    bg_color_frame = "white" if ventana_login.fondo_img else FONDO

    login_frame = tk.Frame(canvas_login, bg=bg_color_frame, bd=2, relief="ridge", padx=20, pady=20)

    tk.Label(login_frame, text="Nombre de usuario", bg=bg_color_frame, font=FUENTE_GRANDE).pack(pady=10)
    usuario_entry = tk.Entry(login_frame, font=FUENTE_GRANDE, width=25)
    usuario_entry.pack(pady=5, padx=20)

    tk.Label(login_frame, text="Contraseña", bg=bg_color_frame, font=FUENTE_GRANDE).pack(pady=10)
    contrasena_entry = tk.Entry(login_frame, show="*", font=FUENTE_GRANDE, width=25)
    contrasena_entry.pack(pady=5, padx=20)

    def iniciar_sesion():
        usuario = usuario_entry.get()
        contrasena = contrasena_entry.get()

        if usuario and contrasena:
            usuarios = cargar_usuarios()
            for u in usuarios:
                if len(u) >= 4 and u[1] == usuario and u[2] == contrasena:
                    global rol_global
                    rol_global = u[3]
                    messagebox.showinfo("Inicio de sesión exitoso", f"Bienvenido, {usuario}!")
                    ventana_login.destroy()
                    abrir_punto_de_venta()
                    return
            messagebox.showerror("Error", "Usuario o contraseña incorrectos.")
        else:
            messagebox.showwarning("Campos vacíos", "Por favor completa todos los campos.")

    tk.Button(login_frame, text="Iniciar Sesión", command=iniciar_sesion, font=FUENTE_GRANDE, width=15).pack(pady=20)

    canvas_login.create_window(ancho_ventana // 2, alto_ventana // 2, window=login_frame, anchor="center")

def abrir_registro():
    # Se pasa la ruta de la imagen para que la ventana de registro tenga el mismo fondo
    ventana_registro, canvas_registro = crear_ventana_con_fondo("Registro de Usuario", ancho=800, alto=600, image_path=r"C:\Users\yorwar\Downloads\tarea de python\imagen de aviones.jpg")
    ancho_ventana, alto_ventana = 800, 600

    bg_color_frame = "white" if ventana_registro.fondo_img else FONDO

    registro_frame = tk.Frame(canvas_registro, bg=bg_color_frame, bd=2, relief="ridge", padx=20, pady=20)

    tk.Label(registro_frame, text="Nombre de usuario", bg=bg_color_frame, font=FUENTE_GRANDE).pack(pady=10)
    usuario_entry = tk.Entry(registro_frame, font=FUENTE_GRANDE, width=25)
    usuario_entry.pack(pady=5, padx=20)

    tk.Label(registro_frame, text="Contraseña", bg=bg_color_frame, font=FUENTE_GRANDE).pack(pady=10)
    contrasena_entry = tk.Entry(registro_frame, show="*", font=FUENTE_GRANDE, width=25)
    contrasena_entry.pack(pady=5, padx=20)

    tk.Label(registro_frame, text="Código de registro", bg=bg_color_frame, font=FUENTE_GRANDE).pack(pady=10)
    codigo_entry = tk.Entry(registro_frame, font=FUENTE_GRANDE, width=25)
    codigo_entry.pack(pady=5, padx=20)

    tk.Label(registro_frame, text="Rol de usuario", bg=bg_color_frame, font=FUENTE_GRANDE).pack(pady=10)
    rol_var = tk.StringVar()
    roles = ["Diseñador", "Analista", "Programador", "Gestor de BD"]
    rol_menu = ttk.Combobox(registro_frame, textvariable=rol_var, values=roles, state="readonly", font=FUENTE_GRANDE, width=23)
    rol_menu.pack(pady=5, padx=20)

    def registrar():
        usuario = usuario_entry.get()
        contrasena = contrasena_entry.get()
        codigo = codigo_entry.get()
        rol = rol_var.get()

        if usuario and contrasena and codigo and rol:
            if codigo == "7448":
                usuarios = cargar_usuarios()
                nuevo_id = max([int(u[0]) for u in usuarios if len(u) > 0 and u[0].isdigit()] + [0]) + 1
                guardar_usuario([nuevo_id, usuario, contrasena, rol])
                messagebox.showinfo("Registro exitoso", "El usuario ha sido registrado exitosamente.")
                ventana_registro.destroy()
            else: 
                messagebox.showerror("Código inválido", "El código de registro es incorrecto.")
        else:
            messagebox.showwarning("Campos vacíos", "Por favor completa todos los campos.")

    tk.Button(registro_frame, text="Registrar", command=registrar, font=FUENTE_GRANDE, width=15).pack(pady=20)

    canvas_registro.create_window(ancho_ventana // 2, alto_ventana // 2, window=registro_frame, anchor="center")

# =====================================================================================
# 6. PUNTO DE ENTRADA PRINCIPAL
# =====================================================================================
if __name__ == "__main__":
    # Crear usuario admin si no existe
    if not os.path.exists(archivo_usuarios) or not any(len(u) >= 2 and u[1] == "admin" for u in cargar_usuarios()):
        usuarios_existentes = cargar_usuarios()
        nuevo_id = len(usuarios_existentes) + 1
        guardar_usuario([nuevo_id, "admin", "admin", "Administrador"])

    # Pantalla principal
    root = tk.Tk()
    root.title("Sistema Punto de  ventas")
    
    ancho_ventana = 1200
    alto_ventana = 800
    root.geometry(f"{ancho_ventana}x{alto_ventana}")

    # --- Cargar imagen de fondo ---
    fondo_img = None
    if Image is not None and ImageTk is not None:
        try:
            # Se actualiza la ruta de la imagen de fondo principal
            image_path = r"C:\Users\yorwar\Downloads\tarea de python\imagen de aviones.jpg"
            if os.path.exists(image_path):
                img = Image.open(image_path)
                img = img.resize((ancho_ventana, alto_ventana), Image.LANCZOS)
                fondo_img = ImageTk.PhotoImage(img)
        except Exception as e:
            print(f"No se pudo cargar la imagen de fondo principal: {e}")
            fondo_img = None

    canvas = tk.Canvas(root, width=ancho_ventana, height=alto_ventana)
    canvas.pack(fill="both", expand=True)

    if fondo_img:
        canvas.create_image(0, 0, image=fondo_img, anchor="nw")
        bg_color = "white"
    else:
        canvas.configure(bg=FONDO)
        bg_color_elementos = FONDO

    bg_color = "white"
    main_frame = tk.Frame(root, bg=bg_color, bd=2, relief="ridge")
    
    tk.Label(main_frame, text="Sistema punto de ventas", font=("Arial", 22, "bold"), bg=main_frame.cget('bg')).pack(pady=20)

    btn_font = ("Arial", 12)
    btn_width = 25
    tk.Button(main_frame, text="Iniciar Sesión", width=btn_width, font=btn_font, command=abrir_login).pack(pady=10)
    tk.Button(main_frame, text="Registro de Usuario", width=btn_width, font=btn_font, command=abrir_registro).pack(pady=10)

    tk.Frame(main_frame, bg=main_frame.cget('bg'), height=30).pack(fill="x") # Espacio

    tk.Label(main_frame, text="Dirección: las flores - Celular: +58 04241796801 Email: yorwarvillamizar@gmail.com ",
                            bg=main_frame.cget('bg'), font=FUENTE).pack()
    tk.Label(main_frame, text="Software creado por yorwar mauricio yohan / InnovadoSolución.8998", bg=main_frame.cget('bg'), font=FUENTE).pack()

    fecha_hora_label = tk.Label(main_frame, bg=main_frame.cget('bg'), font=FUENTE)
    fecha_hora_label.pack(pady=20)
    actualizar_fecha_hora(fecha_hora_label)

    canvas.create_window(ancho_ventana // 2, alto_ventana // 2, window=main_frame, anchor="center")

    # Asignar la imagen al objeto root para mantener la referencia
    if fondo_img:
        root.fondo_img = fondo_img

    root.mainloop()
    